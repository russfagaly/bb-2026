"""
compile.py
==========
Master build script for 2026 Alameda Little League Majors stats workbook.

Usage
-----
    python3 pipeline/compile.py

Reads every game file from games/, aggregates hitting and pitching data,
and writes '2026 ALL MAJORS STATS.xlsx' to the Stats folder.

Always run validate.py first to ensure clean data:
    python3 pipeline/validate.py && python3 pipeline/compile.py

Output tabs
-----------
  1. 🏆 Hitting Leaders   – 18 categories in 3×6 grid
  2. ⚾ Pitching Leaders  – 18 categories in 3×6 grid
  3. 📊 Team Hitting      – per-team hitting totals + AVG/OBP/SLG
  4. 📊 Team Pitching     – per-team pitching totals + ERA/WHIP/Strike%
  5. 🥎 Player Hitting    – individual hitting with Python-computed rate stats
  6. 🎯 Player Pitching   – individual pitching stats
"""

import os, sys, re, importlib.util
from collections import defaultdict

THIS_DIR  = os.path.dirname(os.path.abspath(__file__))
STATS_DIR = os.path.dirname(THIS_DIR)
GAMES_DIR = os.path.join(STATS_DIR, "games")
OUT_PATH  = os.path.join(STATS_DIR, "2026 ALL MAJORS STATS.xlsx")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

# =============================================================================
# 1. LOAD ALL GAME FILES
# =============================================================================
all_hitting  = []   # list of row-dicts: {name, team, date, ab, r, h, ...}
all_pitching = []   # list of row-dicts: {name, team, date, ip, h, r, er, ...}

game_files = sorted(f for f in os.listdir(GAMES_DIR) if f.endswith('.py'))
if not game_files:
    sys.exit(f"No game files found in {GAMES_DIR}")

for fname in game_files:
    path = os.path.join(GAMES_DIR, fname)
    spec = importlib.util.spec_from_file_location("gf", path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    team = mod.TEAM
    date = mod.DATE

    for row in getattr(mod, 'hitting', []):
        all_hitting.append({**row, 'team': team, 'date': date})

    for row in getattr(mod, 'pitching', []):
        all_pitching.append({**row, 'team': team, 'date': date})

print(f"Loaded {len(game_files)} game files: "
      f"{len(all_hitting)} hitting rows, {len(all_pitching)} pitching rows")

# =============================================================================
# 2. HELPER FUNCTIONS
# =============================================================================
def display_name(raw):
    """'Oliver R #41' → 'Oliver R'"""
    return raw.split(' #')[0].strip()

def ip_to_dec(ip_str):
    """'2.1' → 2.333,  '2.2' → 2.667,  '3.0' → 3.0"""
    s = str(ip_str)
    if '.' not in s:
        return float(s)
    whole, frac = s.split('.')
    return int(whole) + int(frac) / 3.0

def dec_to_ip_display(dec):
    """6.333… → '6.1',  6.667… → '6.2'"""
    whole  = int(dec)
    thirds = 0 if (dec - whole) < 0.01 else (1 if (dec - whole) < 0.5 else 2)
    return f"{whole}.{thirds}"

# =============================================================================
# 3. AGGREGATE HITTING  by (display_name, team)
# =============================================================================
h_totals = defaultdict(lambda: {
    'team': '', 'games': 0,
    'ab': 0, 'r': 0, 'h': 0, 'rbi': 0, 'bb': 0, 'so': 0,
    'sb': 0, 'cs': 0, 'e': 0, 'doubles': 0, 'triples': 0, 'hr': 0
})

for row in all_hitting:
    key = (display_name(row['name']), row['team'])
    p   = h_totals[key]
    p['team'] = row['team']
    p['games'] += 1
    for stat in ('ab','r','h','rbi','bb','so','sb','cs','e','doubles','triples','hr'):
        p[stat] += row.get(stat, 0)

# =============================================================================
# 4. AGGREGATE PITCHING  by (display_name, team)
# =============================================================================
p_totals = defaultdict(lambda: {
    'team': '', 'games': 0,
    'ip_dec': 0.0, 'h': 0, 'r': 0, 'er': 0,
    'bb': 0, 'so': 0, 'pitches': 0, 'strikes': 0, 'bf': 0, 'hbp': 0
})

for row in all_pitching:
    key = (display_name(row['name']), row['team'])
    p   = p_totals[key]
    p['team'] = row['team']
    p['games'] += 1
    p['ip_dec'] += ip_to_dec(row['ip'])
    for stat in ('h','r','er','bb','so','hbp'):
        p[stat] += row.get(stat, 0)
    p['pitches'] += row.get('pitches', 0)
    p['strikes'] += row.get('strikes', 0)
    # bf=0 means illegible in source image; skip to avoid pulling down totals
    bf = row.get('bf', 0)
    if bf > 0:
        p['bf'] += bf

# =============================================================================
# 5. TEAM AGGREGATES
# =============================================================================
TEAMS = ['Astros','Brewers','Giants','Guardians','Marlins','Padres','White Sox','Yankees']

team_hit = {t: defaultdict(int) for t in TEAMS}
team_hit_dates = {t: set() for t in TEAMS}

for row in all_hitting:
    t = row['team']
    if t not in team_hit: continue
    team_hit_dates[t].add(row['date'])
    for stat in ('ab','r','h','rbi','bb','so','sb','cs','e','doubles','triples','hr'):
        team_hit[t][stat] += row.get(stat, 0)

team_pit = {t: {'ip_dec': 0.0, 'h': 0, 'r': 0, 'er': 0,
                'bb': 0, 'so': 0, 'pitches': 0, 'strikes': 0,
                'bf': 0, 'hbp': 0} for t in TEAMS}
team_pit_dates = {t: set() for t in TEAMS}

for row in all_pitching:
    t = row['team']
    if t not in team_pit: continue
    team_pit_dates[t].add(row['date'])
    team_pit[t]['ip_dec'] += ip_to_dec(row['ip'])
    for stat in ('h','r','er','bb','so','hbp'):
        team_pit[t][stat] += row.get(stat, 0)
    team_pit[t]['pitches'] += row.get('pitches', 0)
    team_pit[t]['strikes'] += row.get('strikes', 0)
    bf = row.get('bf', 0)
    if bf > 0:
        team_pit[t]['bf'] += bf

# =============================================================================
# 6. STAT FUNCTIONS
# =============================================================================
# Hitting
def h_avg(p):      return p['h'] / p['ab']                      if p['ab'] >= 1          else 0
def h_obp(p):
    d = p['ab'] + p['bb']
    return (p['h'] + p['bb']) / d                                if d >= 1               else 0
def h_slg(p):
    if p['ab'] < 1: return 0
    return (p['h'] + p['doubles'] + 2*p['triples'] + 3*p['hr']) / p['ab']
def h_ops(p):      return h_obp(p) + h_slg(p)
def h_tb(p):       return p['h'] + p['doubles'] + 2*p['triples'] + 3*p['hr']
def h_xbh(p):      return p['doubles'] + p['triples'] + p['hr']
def h_sb_pct(p):   return p['sb'] / (p['sb']+p['cs'])           if (p['sb']+p['cs']) > 0 else 0
def h_walk_pct(p): return p['bb'] / (p['ab']+p['bb'])           if (p['ab']+p['bb']) > 0 else 0
def h_so_pct(p):   return p['so'] / (p['ab']+p['bb'])           if (p['ab']+p['bb']) > 0 else 0

# Pitching
def p_era(p):         return p['er'] * 9 / p['ip_dec']           if p['ip_dec'] > 0.01 else 99.0
def p_whip(p):        return (p['h']+p['bb']) / p['ip_dec']      if p['ip_dec'] > 0.01 else 99.0
def p_strike_pct(p):  return p['strikes'] / p['pitches']         if p['pitches'] > 0   else 0
def p_hits_per_6(p):  return 6 * p['h']  / p['ip_dec']          if p['ip_dec'] > 0     else 0
def p_bb_per_6(p):    return 6 * p['bb'] / p['ip_dec']          if p['ip_dec'] > 0     else 0
def p_so_per_6(p):    return 6 * p['so'] / p['ip_dec']          if p['ip_dec'] > 0     else 0
def p_walk_pct(p):    return p['bb'] / p['bf']                   if p['bf'] > 0         else 0
def p_so_pct(p):      return p['so'] / p['bf']                   if p['bf'] > 0         else 0
def p_k_to_bb(p):     return p['so'] / p['bb']                   if p['bb'] > 0         else 0

# =============================================================================
# 7. PLAYER LISTS  (sorted by team then name for display)
# =============================================================================
hitting_players = sorted(
    [(dname, p) for (dname, _team), p in h_totals.items()],
    key=lambda x: (x[1]['team'], x[0])
)
pitching_players = sorted(
    [(dname, p) for (dname, _team), p in p_totals.items()],
    key=lambda x: (x[1]['team'], x[0])
)

# Leaderboard filter pools
qual_hitters  = [(n, p) for n, p in hitting_players if p['ab'] >= 10]
all_hitters   = hitting_players
sb_hitters    = [(n, p) for n, p in hitting_players if p['sb']+p['cs'] >= 1]

qual_pitchers = [(n, p) for n, p in pitching_players if p['ip_dec'] >= 6.0]
all_pitchers  = pitching_players
kbb_pitchers  = [(n, p) for n, p in pitching_players if p['bb'] > 0 and p['ip_dec'] >= 3.0]

# League averages for OPS+ and ERA+
_lg_h   = sum(p['h']  for _, p in hitting_players)
_lg_bb  = sum(p['bb'] for _, p in hitting_players)
_lg_ab  = sum(p['ab'] for _, p in hitting_players)
_lg_tb  = sum(p['h'] + p['doubles'] + 2*p['triples'] + 3*p['hr'] for _, p in hitting_players)
_lg_obp = (_lg_h + _lg_bb) / (_lg_ab + _lg_bb) if (_lg_ab + _lg_bb) > 0 else 1
_lg_slg = _lg_tb / _lg_ab                       if _lg_ab > 0           else 1
_lg_er     = sum(p['er']     for _, p in pitching_players)
_lg_ip     = sum(p['ip_dec'] for _, p in pitching_players)
_lg_era    = _lg_er * 9 / _lg_ip                if _lg_ip > 0           else 0
_lg_bb_pit = sum(p['bb']     for _, p in pitching_players)
_lg_so_pit = sum(p['so']     for _, p in pitching_players)
_fip_const = _lg_era - (3*_lg_bb_pit - 2*_lg_so_pit) / _lg_ip if _lg_ip > 0 else 0

def h_ops_plus(p):
    obp = (p['h']+p['bb']) / (p['ab']+p['bb']) if (p['ab']+p['bb']) > 0 else 0
    slg = (p['h']+p['doubles']+2*p['triples']+3*p['hr']) / p['ab'] if p['ab'] > 0 else 0
    return round(100 * (obp / _lg_obp + slg / _lg_slg - 1)) if _lg_obp > 0 and _lg_slg > 0 else 100

def p_era_plus(p):
    if p['ip_dec'] <= 0.01: return 0
    era = p['er'] * 9 / p['ip_dec']
    if era == 0:  return 999
    return round(100 * _lg_era / era) if _lg_era > 0 else 0

def p_fip_minus(p):
    if p['ip_dec'] <= 0.01: return 0.0
    return round((3*p['bb'] - 2*p['so']) / p['ip_dec'] + _fip_const, 2)

def p_kbb_ratio(p):
    if p['bb'] == 0: return 99.0 if p['so'] > 0 else 0.0
    return round(p['so'] / p['bb'], 2)

# =============================================================================
# 8. EXCEL STYLE CONSTANTS
# =============================================================================
BLUE_HEADER  = PatternFill('solid', fgColor='1F4E79')
BLUE_SUBHDR  = PatternFill('solid', fgColor='2E75B6')
BLUE_ROW1    = PatternFill('solid', fgColor='D6E4F0')
BLUE_ROW2    = PatternFill('solid', fgColor='FFFFFF')

GREEN_HEADER = PatternFill('solid', fgColor='1E4620')
GREEN_SUBHDR = PatternFill('solid', fgColor='375623')
GREEN_ROW1   = PatternFill('solid', fgColor='D9EAD3')
GREEN_ROW2   = PatternFill('solid', fgColor='FFFFFF')

GRAY_HEADER  = PatternFill('solid', fgColor='404040')
GRAY_SUBHDR  = PatternFill('solid', fgColor='595959')
GRAY_ROW1    = PatternFill('solid', fgColor='F2F2F2')
GRAY_ROW2    = PatternFill('solid', fgColor='FFFFFF')

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_title(ws, row, c1, c2, text, font, fill):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font      = font
    c.fill      = fill
    c.alignment = CENTER
    return c

# =============================================================================
# 9. LEADERBOARD GRID
# =============================================================================
TOP_N        = 10
BLOCK_HEIGHT = TOP_N + 2   # title + col-header + 10 data rows = 12
ROW_GAP      = 2           # blank rows between grid rows
BLOCK_COLS   = [1, 6, 11, 16, 21, 26]   # 6 blocks across

def write_lb_block(ws, row, col, title, pool, sort_fn, fmt, higher_better,
                   subhdr_fill, row1_fill, row2_fill, display_fn=None):
    """Write a 4-column leaderboard block (Rk, Player, Team, Stat)."""
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+3)
    c = ws.cell(row=row, column=col, value=title)
    c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    c.fill      = subhdr_fill
    c.alignment = CENTER
    ws.row_dimensions[row].height = 17

    for i, h in enumerate(['Rk', 'Player', 'Team', 'Stat']):
        cc = ws.cell(row=row+1, column=col+i, value=h)
        cc.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cc.fill      = subhdr_fill
        cc.alignment = CENTER
    ws.row_dimensions[row+1].height = 13

    sorted_pool = sorted(pool, key=lambda x: sort_fn(x[1]), reverse=higher_better)[:TOP_N]

    for rank, (name, p) in enumerate(sorted_pool, 1):
        dr   = row + 1 + rank
        fill = row1_fill if rank % 2 == 1 else row2_fill
        raw  = sort_fn(p)
        disp = display_fn(raw) if display_fn else raw

        for i, v in enumerate([rank, name, p['team'], disp]):
            cc = ws.cell(row=dr, column=col+i, value=v)
            cc.fill      = fill
            cc.font      = Font(name='Arial', size=9)
            cc.alignment = LEFT if i == 1 else CENTER
            if i == 3 and fmt and display_fn is None:
                cc.number_format = fmt
        ws.row_dimensions[dr].height = 14


def write_lb_grid(ws, start_row, categories, subhdr_fill, row1_fill, row2_fill):
    """Write 18 leaderboard blocks in a 3-row × 6-column grid."""
    r = start_row
    for row_idx in range(3):
        for col_idx in range(6):
            cat_idx = row_idx * 6 + col_idx
            if cat_idx >= len(categories): break
            cat = categories[cat_idx]
            title, pool, sort_fn, fmt, higher_better = cat[:5]
            display_fn = cat[5] if len(cat) > 5 else None
            write_lb_block(ws, r, BLOCK_COLS[col_idx],
                           title, pool, sort_fn, fmt, higher_better,
                           subhdr_fill, row1_fill, row2_fill, display_fn=display_fn)
        r += BLOCK_HEIGHT + ROW_GAP


def set_lb_col_widths(ws):
    for col_start in BLOCK_COLS:
        col_width(ws, col_start,   4)
        col_width(ws, col_start+1, 15)
        col_width(ws, col_start+2, 11)
        col_width(ws, col_start+3, 8)
        if col_start < BLOCK_COLS[-1]:
            col_width(ws, col_start+4, 2)

# =============================================================================
# 10. BUILD WORKBOOK
# =============================================================================
wb = Workbook()
wb.remove(wb.active)

# ── TAB 1: Hitting Leaders ────────────────────────────────────────────────────
ws_hl = wb.create_sheet("🏆 Hitting Leaders")
ws_hl.sheet_properties.tabColor = "2E75B6"

merge_title(ws_hl, 1, 1, 29,
            "2026 Alameda Little League Majors — HITTING LEADERBOARD",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), BLUE_HEADER)
ws_hl.row_dimensions[1].height = 30

merge_title(ws_hl, 2, 1, 29,
            "Min 10 AB for rate stats (AVG/OBP/SLG/OPS/Walk%/K%).  "
            "Min 1 SB attempt for SB%.",
            Font(name='Arial', italic=True, size=9, color='FFFFFF'), BLUE_SUBHDR)
ws_hl.row_dimensions[2].height = 15

HIT_CATS = [
    # Row 1
    ("Batting Avg (AVG)",  qual_hitters, h_avg,                  '0.000', True),
    ("On-Base Pct (OBP)",  qual_hitters, h_obp,                  '0.000', True),
    ("Slugging Pct (SLG)", qual_hitters, h_slg,                  '0.000', True),
    ("OPS",                qual_hitters, h_ops,                  '0.000', True),
    ("Hits (H)",           all_hitters,  lambda p: p['h'],        None,    True),
    ("RBI",                all_hitters,  lambda p: p['rbi'],      None,    True),
    # Row 2
    ("Runs Scored (R)",    all_hitters,  lambda p: p['r'],        None,    True),
    ("Total Bases (TB)",   all_hitters,  h_tb,                    None,    True),
    ("Walks (BB)",         all_hitters,  lambda p: p['bb'],       None,    True),
    ("Doubles (2B)",       all_hitters,  lambda p: p['doubles'],  None,    True),
    ("Triples (3B)",       all_hitters,  lambda p: p['triples'],  None,    True),
    ("Home Runs (HR)",     all_hitters,  lambda p: p['hr'],       None,    True),
    # Row 3
    ("Stolen Bases (SB)",  all_hitters,  lambda p: p['sb'],       None,    True),
    ("SB% (min 1 att)",    sb_hitters,   h_sb_pct,                '0.0%',  True),
    ("Walk % (min 10 AB)", qual_hitters, h_walk_pct,              '0.0%',  True),
    ("K% Fewest (min 10)", qual_hitters, h_so_pct,                '0.0%',  False),
    ("Extra Base Hits",    all_hitters,  h_xbh,                   None,    True),
    ("Errors (Most)",      all_hitters,  lambda p: p['e'],        None,    True),
]

write_lb_grid(ws_hl, 4, HIT_CATS, BLUE_SUBHDR, BLUE_ROW1, BLUE_ROW2)
set_lb_col_widths(ws_hl)

# ── TAB 2: Pitching Leaders ───────────────────────────────────────────────────
ws_pl = wb.create_sheet("⚾ Pitching Leaders")
ws_pl.sheet_properties.tabColor = "375623"

merge_title(ws_pl, 1, 1, 29,
            "2026 Alameda Little League Majors — PITCHING LEADERBOARD",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), GREEN_HEADER)
ws_pl.row_dimensions[1].height = 30

merge_title(ws_pl, 2, 1, 29,
            "Min 3 IP for rate stats (ERA/WHIP/Strike%/H·BB·K per 6/Walk%/K%).  "
            "Min 1 BB for K/BB ratio.",
            Font(name='Arial', italic=True, size=9, color='FFFFFF'), GREEN_SUBHDR)
ws_pl.row_dimensions[2].height = 15

PIT_CATS = [
    # Row 1
    ("Innings Pitched",        all_pitchers,  lambda p: p['ip_dec'],  None,   True,
     dec_to_ip_display),
    ("ERA (min 6 IP)",         qual_pitchers, p_era,                  '0.00', False),
    ("Hits Allowed (Fewest, min 6 IP)",  qual_pitchers,  lambda p: p['h'],       None,   False),
    ("Runs Allowed (Most)",    all_pitchers,  lambda p: p['r'],       None,   True),
    ("Earned Runs (Most)",     all_pitchers,  lambda p: p['er'],      None,   True),
    ("Walks",                  all_pitchers,  lambda p: p['bb'],      None,   True),
    # Row 2
    ("Strikeouts (SO)",        all_pitchers,  lambda p: p['so'],      None,   True),
    ("Pitches Thrown",         all_pitchers,  lambda p: p['pitches'], None,   True),
    ("Strike % (min 6 IP)",    qual_pitchers, p_strike_pct,           '0.0%', True),
    ("HBP Most",               all_pitchers,  lambda p: p['hbp'],     None,   True),
    ("WHIP (min 6 IP)",        qual_pitchers, p_whip,                 '0.00', False),
    ("H per 6 IP (min 6 IP)",  qual_pitchers, p_hits_per_6,           '0.00', False),
    # Row 3
    ("BB per 6 IP (min 6 IP)", qual_pitchers, p_bb_per_6,             '0.00', False),
    ("K per 6 IP (min 6 IP)",  qual_pitchers, p_so_per_6,             '0.00', True),
    ("Walk % (min 6 IP)",      qual_pitchers, p_walk_pct,             '0.0%', False),
    ("K% (min 6 IP)",          qual_pitchers, p_so_pct,               '0.0%', True),
    ("Batters Faced (BF)",     all_pitchers,  lambda p: p['bf'],      None,   True),
    ("K/BB Ratio (min 6 IP)",  kbb_pitchers,  p_k_to_bb,              '0.00', True),
]

write_lb_grid(ws_pl, 4, PIT_CATS, GREEN_SUBHDR, GREEN_ROW1, GREEN_ROW2)
set_lb_col_widths(ws_pl)

# ── TAB 3: Team Hitting ───────────────────────────────────────────────────────
ws_th = wb.create_sheet("📊 Team Hitting")
ws_th.sheet_properties.tabColor = "595959"

merge_title(ws_th, 1, 1, 17,
            "2026 Alameda Little League Majors — TEAM HITTING STATISTICS",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), GRAY_HEADER)
ws_th.row_dimensions[1].height = 30

r = 3
for col, h in enumerate(['Team','G','AB','R','H','2B','3B','HR','RBI','BB','SO',
                          'SB','CS','E','AVG','OBP','SLG'], 1):
    c = ws_th.cell(row=r, column=col, value=h)
    c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill      = GRAY_SUBHDR
    c.alignment = CENTER
ws_th.row_dimensions[r].height = 18
r += 1

for i, team in enumerate(TEAMS):
    th   = team_hit[team]
    g    = len(team_hit_dates[team])
    fill = GRAY_ROW1 if i % 2 == 0 else GRAY_ROW2

    count_vals = [team, g, th['ab'], th['r'], th['h'],
                  th['doubles'], th['triples'], th['hr'], th['rbi'],
                  th['bb'], th['so'], th['sb'], th['cs'], th['e']]
    for col, v in enumerate(count_vals, 1):
        c = ws_th.cell(row=r, column=col, value=v)
        c.fill      = fill
        c.font      = Font(name='Arial', size=10)
        c.alignment = LEFT if col == 1 else CENTER

    ab  = th['ab'];  h_ = th['h'];  bb = th['bb']
    d   = th['doubles'];  t_ = th['triples'];  hr = th['hr']
    avg = h_ / ab                              if ab > 0         else 0
    obp = (h_ + bb) / (ab + bb)               if (ab + bb) > 0  else 0
    slg = (h_ + d + 2*t_ + 3*hr) / ab         if ab > 0         else 0

    for col, val, fmt in [(15, avg, '0.000'), (16, obp, '0.000'), (17, slg, '0.000')]:
        c = ws_th.cell(row=r, column=col, value=val)
        c.fill         = fill
        c.font         = Font(name='Arial', size=10)
        c.alignment    = CENTER
        c.number_format = fmt

    ws_th.row_dimensions[r].height = 16
    r += 1

for col, w in [(1,14),(2,5),(3,7),(4,6),(5,6),(6,6),(7,6),(8,6),(9,6),
               (10,6),(11,6),(12,6),(13,6),(14,6),(15,8),(16,8),(17,8)]:
    col_width(ws_th, col, w)

ws_th.auto_filter.ref = "A3:Q3"
ws_th.freeze_panes    = "A4"

# ── TAB 4: Team Pitching ──────────────────────────────────────────────────────
ws_tp = wb.create_sheet("📊 Team Pitching")
ws_tp.sheet_properties.tabColor = "595959"

merge_title(ws_tp, 1, 1, 16,
            "2026 Alameda Little League Majors — TEAM PITCHING STATISTICS",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), GRAY_HEADER)
ws_tp.row_dimensions[1].height = 30

r = 3
for col, h in enumerate(['Team','G','IP','H','R','ER','BB','SO','HBP','BF',
                          'ERA','WHIP','K/IP','Pitches','Strikes','Strike%'], 1):
    c = ws_tp.cell(row=r, column=col, value=h)
    c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill      = GRAY_SUBHDR
    c.alignment = CENTER
ws_tp.row_dimensions[r].height = 18
r += 1

for i, team in enumerate(TEAMS):
    tp   = team_pit[team]
    g    = len(team_pit_dates[team])
    fill = GRAY_ROW1 if i % 2 == 0 else GRAY_ROW2
    ip_d = tp['ip_dec']

    count_vals = [team, g, dec_to_ip_display(ip_d),
                  tp['h'], tp['r'], tp['er'], tp['bb'], tp['so'],
                  tp['hbp'], tp['bf']]
    for col, v in enumerate(count_vals, 1):
        c = ws_tp.cell(row=r, column=col, value=v)
        c.fill      = fill
        c.font      = Font(name='Arial', size=10)
        c.alignment = LEFT if col == 1 else CENTER

    era        = round(tp['er'] * 9 / ip_d,          2) if ip_d > 0 else 0.0
    whip       = round((tp['h'] + tp['bb']) / ip_d,  2) if ip_d > 0 else 0.0
    kip        = round(tp['so'] / ip_d,              2) if ip_d > 0 else 0.0
    strike_pct = tp['strikes'] / tp['pitches']           if tp['pitches'] > 0 else 0.0

    for col, val, fmt in [
        (11, era,           '0.00'),
        (12, whip,          '0.00'),
        (13, kip,           '0.00'),
        (14, tp['pitches'], '0'),
        (15, tp['strikes'], '0'),
        (16, strike_pct,    '0.0%'),
    ]:
        c = ws_tp.cell(row=r, column=col, value=val)
        c.fill         = fill
        c.font         = Font(name='Arial', size=10)
        c.alignment    = CENTER
        c.number_format = fmt

    ws_tp.row_dimensions[r].height = 16
    r += 1

for col, w in [(1,14),(2,5),(3,7),(4,6),(5,6),(6,6),(7,6),(8,6),(9,6),(10,6),
               (11,7),(12,7),(13,7),(14,9),(15,9),(16,9)]:
    col_width(ws_tp, col, w)

ws_tp.auto_filter.ref = "A3:P3"
ws_tp.freeze_panes    = "A4"

# ── TAB 5: Player Hitting ─────────────────────────────────────────────────────
ws_ph = wb.create_sheet("🥎 Player Hitting")
ws_ph.sheet_properties.tabColor = "2E75B6"

merge_title(ws_ph, 1, 1, 20,
            "2026 Alameda Little League Majors — PLAYER HITTING STATISTICS",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), BLUE_HEADER)
ws_ph.row_dimensions[1].height = 30

r = 3
for col, h in enumerate(['Player','Team','GP','AB','R','H','2B','3B','HR','RBI',
                          'BB','SO','SB','CS','E','AVG','OBP','SLG','OPS','OPS+'], 1):
    c = ws_ph.cell(row=r, column=col, value=h)
    c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill      = BLUE_SUBHDR
    c.alignment = CENTER
ws_ph.row_dimensions[r].height = 18
r += 1

for i, (name, p) in enumerate(hitting_players):
    fill = BLUE_ROW1 if i % 2 == 0 else BLUE_ROW2
    ab   = p['ab'];  h_ = p['h'];  bb = p['bb']
    d    = p['doubles'];  t_ = p['triples'];  hr = p['hr']

    avg = h_ / ab                      if ab > 0        else 0
    obp = (h_ + bb) / (ab + bb)        if (ab + bb) > 0 else 0
    slg = (h_ + d + 2*t_ + 3*hr) / ab if ab > 0        else 0
    ops = obp + slg

    count_vals = [name, p['team'], p['games'], ab, p['r'], h_,
                  d, t_, hr, p['rbi'], bb, p['so'], p['sb'], p['cs'], p['e']]
    for col, v in enumerate(count_vals, 1):
        c = ws_ph.cell(row=r, column=col, value=v)
        c.fill      = fill
        c.font      = Font(name='Arial', size=10)
        c.alignment = LEFT if col <= 2 else CENTER

    ops_plus = h_ops_plus(p)
    for col, val, fmt in [(16, avg,'0.000'),(17, obp,'0.000'),
                           (18, slg,'0.000'),(19, ops,'0.000'),(20, ops_plus,'0')]:
        c = ws_ph.cell(row=r, column=col, value=val)
        c.fill         = fill
        c.font         = Font(name='Arial', size=10)
        c.alignment    = CENTER
        c.number_format = fmt

    ws_ph.row_dimensions[r].height = 15
    r += 1

for col, w in [(1,18),(2,12),(3,5),(4,5),(5,5),(6,5),(7,5),(8,5),(9,5),
               (10,5),(11,5),(12,5),(13,5),(14,5),(15,5),
               (16,8),(17,8),(18,8),(19,8),(20,8)]:
    col_width(ws_ph, col, w)

ws_ph.auto_filter.ref = "A3:T3"   # dropdown arrows on every column header
ws_ph.freeze_panes    = "A4"       # keep title + header row visible while scrolling

# ── TAB 6: Player Pitching ────────────────────────────────────────────────────
ws_pp = wb.create_sheet("🎯 Player Pitching")
ws_pp.sheet_properties.tabColor = "375623"

merge_title(ws_pp, 1, 1, 20,
            "2026 Alameda Little League Majors — PLAYER PITCHING STATISTICS",
            Font(name='Arial', bold=True, size=16, color='FFFFFF'), GREEN_HEADER)
ws_pp.row_dimensions[1].height = 30

r = 3
for col, h in enumerate(['Player','Team','G','IP','H','R','ER','BB','SO','HBP',
                          'Pitches','Strikes','Strike%','BF','ERA','WHIP','K/IP','K/BB','FIP-','ERA+'], 1):
    c = ws_pp.cell(row=r, column=col, value=h)
    c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill      = GREEN_SUBHDR
    c.alignment = CENTER
ws_pp.row_dimensions[r].height = 18
r += 1

for i, (name, p) in enumerate(pitching_players):
    fill   = GREEN_ROW1 if i % 2 == 0 else GREEN_ROW2
    ip_d   = p['ip_dec']
    str_pc = p['strikes'] / p['pitches'] if p['pitches'] > 0 else 0.0
    era    = round(p['er'] * 9 / ip_d,        2) if ip_d > 0 else 0.0
    whip   = round((p['h'] + p['bb']) / ip_d, 2) if ip_d > 0 else 0.0
    kip    = round(p['so'] / ip_d,            2) if ip_d > 0 else 0.0

    count_vals = [name, p['team'], p['games'], dec_to_ip_display(ip_d),
                  p['h'], p['r'], p['er'], p['bb'], p['so'], p['hbp'],
                  p['pitches'], p['strikes']]
    for col, v in enumerate(count_vals, 1):
        c = ws_pp.cell(row=r, column=col, value=v)
        c.fill      = fill
        c.font      = Font(name='Arial', size=10)
        c.alignment = LEFT if col <= 2 else CENTER

    era_plus  = p_era_plus(p)
    fip_minus = p_fip_minus(p)
    kbb       = p_kbb_ratio(p)
    for col, val, fmt in [
        (13, str_pc,   '0.0%'),
        (14, p['bf'],  '0'),
        (15, era,      '0.00'),
        (16, whip,     '0.00'),
        (17, kip,      '0.00'),
        (18, kbb,      '0.00'),
        (19, fip_minus,'0.00'),
        (20, era_plus, '0'),
    ]:
        c = ws_pp.cell(row=r, column=col, value=val)
        c.fill         = fill
        c.font         = Font(name='Arial', size=10)
        c.alignment    = CENTER
        c.number_format = fmt

    ws_pp.row_dimensions[r].height = 15
    r += 1

for col, w in [(1,18),(2,12),(3,5),(4,6),(5,5),(6,5),(7,5),(8,5),(9,5),
               (10,5),(11,8),(12,8),(13,8),(14,5),(15,8),(16,8),(17,8),(18,8),(19,8),(20,8)]:
    col_width(ws_pp, col, w)

ws_pp.auto_filter.ref = "A3:T3"
ws_pp.freeze_panes    = "A4"

# =============================================================================
# 11. SAVE
# =============================================================================
wb.save(OUT_PATH)
print(f"\n✅  Saved: {OUT_PATH}")
print(f"    Tabs:  {wb.sheetnames}")

# Quick summary
print(f"\nSummary:")
print(f"  Players (hitting):  {len(hitting_players)}")
print(f"  Players (pitching): {len(pitching_players)}")
print(f"  Teams:              {len(TEAMS)}")
print(f"  Games loaded:       {len(game_files)}")
